"""
Retail Crisis & Recovery — Hidden Growth Product Detection
DQLab x UjiKompetensi Hackathon | HACK-2026-PYTHON-01

Mengidentifikasi produk "rising star" yang tumbuh konsisten namun tidak
terlihat di agregasi tradisional (Top N), lalu mencari pola pembelian
bersama untuk strategi bundling menggunakan Apriori.

Output:
    retail_insight.xlsx     — Sheet Rising Star + Potential Packaging
    rising_star_index.png   — Line chart pertumbuhan relatif (Base 100)
    rising_star_actual.png  — Line chart nilai penjualan aktual
"""

import os
import sys

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from mlxtend.frequent_patterns import apriori, association_rules
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# HELPER — EXCEL STYLING
# ─────────────────────────────────────────────────────────────────────────────

def thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


# ─────────────────────────────────────────────────────────────────────────────
# HELPER — RISING STAR: Streak Detection
# ─────────────────────────────────────────────────────────────────────────────

def calc_max_streak(group: pd.DataFrame) -> dict:
    """
    Mencari sesi tren naik terpanjang (consecutive rising MA days) per produk.

    Pendekatan:
    - Iterasi nilai numpy array (bukan iterrows) untuk menghindari overhead
      pembuatan Series per baris — ~50-100x lebih cepat untuk dataset besar.
    - Mengembalikan streak terpanjang beserta growth% sesi tersebut.

    Growth % = (MA akhir sesi / MA awal sesi - 1) * 100
    """
    is_rising = group['is_rising'].values
    ma3_vals  = group['ma3'].values

    best_streak = best_start = best_end = 0
    streak = start_idx = 0

    for i, rising in enumerate(is_rising):
        if rising:
            if streak == 0:
                start_idx = i
            streak += 1
            if streak > best_streak:
                best_streak = streak
                best_start  = start_idx
                best_end    = i
        else:
            streak = 0

    if best_streak == 0:
        return {'max_streak': 0, 'growth_pct': 0.0}

    start_ma = ma3_vals[best_start]
    end_ma   = ma3_vals[best_end]
    growth   = (end_ma / start_ma - 1) * 100 if start_ma > 0 else 0.0

    return {'max_streak': best_streak, 'growth_pct': growth}


# ─────────────────────────────────────────────────────────────────────────────
# HELPER — VISUALISASI: Legend Sorting
# ─────────────────────────────────────────────────────────────────────────────

def sort_legend(ax: plt.Axes):
    """Top Sales dulu, lalu Rising Star diurutkan by rank number."""
    handles, labels = ax.get_legend_handles_labels()
    top  = [(h, l) for h, l in zip(handles, labels) if l.startswith('Top Sales')]
    star = [(h, l) for h, l in zip(handles, labels) if not l.startswith('Top Sales')]
    star = sorted(star, key=lambda x: int(x[1].split(':')[0].split()[-1]))
    merged = top + star
    return [x[0] for x in merged], [x[1] for x in merged]


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():

    # ── 0. LOAD & VALIDATE DATA ───────────────────────────────────────────────
    data_file = 'data_penjualan.xlsx'
    if not os.path.exists(data_file):
        sys.exit(f"[ERROR] File '{data_file}' tidak ditemukan di: {os.getcwd()}")

    df = pd.read_excel(data_file, parse_dates=['tgl_transaksi'])

    required_cols = {
        'nomor_struk', 'tgl_transaksi', 'kode_produk',
        'nama_produk', 'jumlah_terjual', 'harga', 'total_nilai'
    }
    missing_cols = required_cols - set(df.columns)
    if missing_cols:
        sys.exit(f"[ERROR] Kolom tidak ditemukan: {missing_cols}")

    print(f"Data loaded  : {df.shape[0]:,} baris")
    print(f"Struk unik   : {df['nomor_struk'].nunique():,}")
    print(f"Produk unik  : {df['kode_produk'].nunique():,}")
    print(f"Rentang tgl  : {df['tgl_transaksi'].min().date()} "
          f"s/d {df['tgl_transaksi'].max().date()}")


    # ── 1. RISING STAR DETECTION ──────────────────────────────────────────────

    # Agregasi nilai penjualan harian per produk
    daily_df = (
        df.groupby(['kode_produk', 'nama_produk', 'tgl_transaksi'], as_index=False)
          ['total_nilai'].sum()
          .sort_values(['kode_produk', 'tgl_transaksi'])
    )

    # Moving Average 3 hari
    # min_periods=3: MA valid hanya setelah ada 3 data point (sesuai spesifikasi)
    daily_df['ma3'] = (
        daily_df
        .groupby('kode_produk')['total_nilai']
        .transform(lambda x: x.rolling(3, min_periods=3).mean())
    )
    daily_df['ma_prev']   = daily_df.groupby('kode_produk')['ma3'].shift(1)
    daily_df['is_rising'] = daily_df['ma3'] > daily_df['ma_prev']

    # Hitung streak terpanjang per produk
    streak_results = (
        daily_df
        .dropna(subset=['ma_prev', 'ma3'])
        .groupby(['kode_produk', 'nama_produk'])
        .apply(calc_max_streak, include_groups=False)
        .apply(pd.Series)
        .reset_index()
    )

    total_sales = (
        df.groupby('kode_produk', as_index=False)['total_nilai']
          .sum()
          .rename(columns={'total_nilai': 'total_penjualan'})
    )

    # Filter: hanya produk dengan max streak >= 12 hari, sort by growth desc
    rising_stars = (
        streak_results[streak_results['max_streak'] >= 12]
        .merge(total_sales, on='kode_produk')
        .sort_values('growth_pct', ascending=False)
        .reset_index(drop=True)
    )

    print(f"\nJumlah Rising Star: {len(rising_stars)}")
    print(rising_stars[
        ['kode_produk', 'nama_produk', 'max_streak', 'growth_pct', 'total_penjualan']
    ].to_string())


    # ── 2. VISUALISASI ────────────────────────────────────────────────────────

    top3_codes = (
        df.groupby('kode_produk')['total_nilai']
          .sum().nlargest(3).index.tolist()
    )
    rs_codes      = rising_stars['kode_produk'].tolist()
    prod_name_map = daily_df.groupby('kode_produk')['nama_produk'].first()
    all_dates     = pd.date_range(daily_df['tgl_transaksi'].min(),
                                  daily_df['tgl_transaksi'].max())

    def prepare_series(sub_df: pd.DataFrame) -> pd.DataFrame:
        """
        Reindex ke rentang tanggal penuh.
        Hari tanpa transaksi = 0 (bukan interpolasi).
        MA dihitung ulang setelah pengisian agar konsisten dengan data harian.
        """
        sub_df = sub_df.set_index('tgl_transaksi').reindex(all_dates)
        sub_df['total_nilai'] = sub_df['total_nilai'].fillna(0)
        sub_df['ma3'] = sub_df['total_nilai'].rolling(3, min_periods=3).mean()
        sub_df.index.name = 'tgl_transaksi'
        return sub_df.reset_index()

    def add_normalized(group: pd.DataFrame) -> pd.DataFrame:
        """Normalisasi MA ke Base 100 dari nilai valid pertama."""
        group     = prepare_series(group.copy())
        valid_pos = group['ma3'].dropna()
        valid_pos = valid_pos[valid_pos > 0]
        first_val = valid_pos.iloc[0] if len(valid_pos) > 0 else np.nan
        group['Normalized'] = (group['ma3'] / first_val * 100) if first_val else np.nan
        return group

    plot_df = pd.concat([
        add_normalized(g)
        for _, g in daily_df[daily_df['kode_produk'].isin(rs_codes)].groupby('kode_produk')
    ])
    plot_df['nama_produk'] = plot_df['kode_produk'].map(prod_name_map)

    top3_plot_df = pd.concat([
        add_normalized(g)
        for _, g in daily_df[daily_df['kode_produk'].isin(top3_codes)].groupby('kode_produk')
    ])
    top3_plot_df['nama_produk'] = top3_plot_df['kode_produk'].map(prod_name_map)

    # Palet warna
    custom_palette = ['#FFD700', '#C0C0C0', '#CD7F32', '#2ecc71',
                      '#3498db', '#9b59b6', '#e74c3c', '#34495e']
    default_color  = '#95a5a6'
    grey_colors    = ['#B0B0B0', '#909090', '#707070']

    color_mapping, rank_mapping = {}, {}
    for i, row in enumerate(rising_stars.itertuples()):
        color_mapping[row.kode_produk] = (
            custom_palette[i] if i < len(custom_palette) else default_color
        )
        
        rank_mapping[row.kode_produk] = i + 1

    font_title = {'family': 'sans-serif', 'color': 'black', 'weight': 'bold', 'size': 16}
    font_label = {'family': 'sans-serif', 'weight': 'normal', 'size': 12}

    def plot_lines(ax, top_df, rs_df, y_col):
        for idx, (kode, grp) in enumerate(top_df.groupby('kode_produk')):
            nama = prod_name_map.get(kode, kode)
            ax.plot(grp['tgl_transaksi'], grp[y_col],
                    linestyle='--', linewidth=2, marker='o', markersize=3,
                    color=grey_colors[idx % len(grey_colors)], alpha=0.7,
                    label=f"Top Sales: {nama}")
        for kode, grp in rs_df.groupby('kode_produk'):
            nama = prod_name_map.get(kode, kode)
            ax.plot(grp['tgl_transaksi'], grp[y_col],
                    marker='o', markersize=4, linewidth=2.5,
                    color=color_mapping.get(kode, default_color),
                    label=f"Rank {rank_mapping.get(kode, '?')}: {nama}")

    # Grafik 1 — Pertumbuhan Relatif (Index Base 100)
    fig, ax = plt.subplots(figsize=(15, 8), dpi=100)
    plot_lines(ax, top3_plot_df, plot_df, 'Normalized')
    ax.set_title('ANALISIS PERTUMBUHAN RELATIF PRODUK RISING STAR\n'
                 '(Dengan Benchmark Top 3 Total Penjualan)',
                 fontdict=font_title, pad=20)
    ax.set_xlabel('Periode Tanggal', fontdict=font_label, labelpad=10)
    ax.set_ylabel('Indeks Pertumbuhan (Base 100)', fontdict=font_label, labelpad=10)
    ax.grid(True, linestyle='--', linewidth=0.5, alpha=0.5)
    ax.axhline(y=100, color='black', linestyle='-', linewidth=1, alpha=0.5)
    plt.xticks(rotation=45, ha='right', fontsize=10)
    plt.yticks(fontsize=10)
    h, l = sort_legend(ax)
    ax.legend(h, l, title="Kategori Produk", title_fontsize=12, fontsize=10,
              bbox_to_anchor=(1.02, 1), loc='upper left',
              borderaxespad=0, frameon=True, shadow=True)
    plt.tight_layout()
    plt.savefig('rising_star_index.png', bbox_inches='tight')
    plt.close()
    print("\nGrafik disimpan: rising_star_index.png")

    # Grafik 2 — Nilai Penjualan Aktual
    fig2, ax2 = plt.subplots(figsize=(15, 8), dpi=100)
    plot_lines(ax2, top3_plot_df, plot_df, 'total_nilai')
    ax2.set_title('ANALISIS NILAI PENJUALAN PRODUK RISING STAR\n(Nilai Penjualan Asli)',
                  fontdict=font_title, pad=20)
    ax2.set_xlabel('Periode Tanggal', fontdict=font_label, labelpad=10)
    ax2.set_ylabel('Total Nilai Penjualan', fontdict=font_label, labelpad=10)
    ax2.grid(True, linestyle='--', linewidth=0.5, alpha=0.5)
    plt.xticks(rotation=45, ha='right', fontsize=10)
    plt.yticks(fontsize=10)
    h2, l2 = sort_legend(ax2)
    ax2.legend(h2, l2, title="Kategori Produk", title_fontsize=12, fontsize=10,
               bbox_to_anchor=(1.02, 1), loc='upper left',
               borderaxespad=0, frameon=True, shadow=True)
    plt.tight_layout()
    plt.savefig('rising_star_actual.png', bbox_inches='tight')
    plt.close()
    print("Grafik disimpan: rising_star_actual.png")


    # ── 3. POTENTIAL PACKAGING — APRIORI (mlxtend) ────────────────────────────

    # Basket matrix: baris = invoice, kolom = produk, nilai = bool (dibeli / tidak)
    basket_matrix = (
        df.groupby(['nomor_struk', 'nama_produk'])['jumlah_terjual']
          .sum()
          .unstack(fill_value=0)
          .astype(bool)
    )

    n_inv = basket_matrix.shape[0]
    print(f"\nMatrix transaksi: {n_inv:,} struk × {basket_matrix.shape[1]:,} produk")

    # Frequent itemsets dengan Apriori (min_support = 1%)
    freq_itemsets = apriori(basket_matrix, min_support=0.01, use_colnames=True)

    # Association rules: metric lift, min_threshold = 1 (sesuai spesifikasi soal)
    rules = association_rules(
        freq_itemsets, metric='lift', min_threshold=1.0,
        num_itemsets=len(freq_itemsets)
    )

    # Filter: setidaknya satu sisi mengandung rising star + lift >= 2
    rs_names = set(rising_stars['nama_produk'])

    def contains_rs(itemset) -> bool:
        return bool(set(itemset) & rs_names)

    mask = (
        rules.apply(
            lambda r: contains_rs(r['antecedents']) or contains_rs(r['consequents']),
            axis=1
        )
        & (rules['lift'] >= 2)
    )
    rules_f = rules[mask].copy()

    # Format kolom output
    rules_f['Jika Membeli']   = rules_f['antecedents'].apply(lambda x: ', '.join(sorted(x)))
    rules_f['Maka Membeli']   = rules_f['consequents'].apply(lambda x: ', '.join(sorted(x)))
    rules_f['Jumlah Invoice'] = (rules_f['support'] * n_inv).round(0).astype(int)
    rules_f['Support']        = rules_f['support'].round(2)
    rules_f['Confidence']     = rules_f['confidence'].round(2)
    rules_f['Lift']           = rules_f['lift'].round(2)

    output_pp = (
        rules_f[['Jika Membeli', 'Maka Membeli', 'Jumlah Invoice',
                  'Support', 'Confidence', 'Lift']]
        .sort_values(['Lift', 'Support', 'Confidence'], ascending=False)
        .reset_index(drop=True)
    )

    print(f"Jumlah rules Potential Packaging: {len(output_pp)}")


    # ── 4. EXPORT KE EXCEL ────────────────────────────────────────────────────

    wb = Workbook()

    # ── Sheet 1: Rising Star ──
    ws1 = wb.active
    ws1.title = "Rising Star"
    ws1.sheet_view.showGridLines = False

    headers_rs   = ["Kode Produk", "Nama Produk", "Growth %", "Total Penjualan"]
    col_widths_rs = [14, 28, 12, 18]
    for ci, (h, w) in enumerate(zip(headers_rs, col_widths_rs), 1):
        c            = ws1.cell(row=1, column=ci, value=h)
        c.font       = Font(bold=True, size=10, name="Arial")
        c.border     = thin_border()
        c.alignment  = Alignment(horizontal="center", vertical="center")
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[1].height = 22

    for ri, row in rising_stars.iterrows():
        r = ri + 2
        cells = [
            (1, row['kode_produk'],          None,    'center'),
            (2, row['nama_produk'],          None,    'left'),
            (3, round(row['growth_pct'], 2), '0.00',  'center'),
            (4, int(row['total_penjualan']), '#,##0', 'right'),
        ]
        for ci, val, fmt, aln in cells:
            c           = ws1.cell(row=r, column=ci, value=val)
            c.font      = Font(size=10, name="Arial")
            c.border    = thin_border()
            c.alignment = Alignment(horizontal=aln)
            if fmt:
                c.number_format = fmt
        ws1.row_dimensions[r].height = 18

    # ── Sheet 2: Potential Packaging ──
    ws2 = wb.create_sheet("Potential Packaging")
    ws2.sheet_view.showGridLines = False

    headers_pp    = ["Jika Membeli", "Maka Membeli", "Jumlah Invoice",
                     "Support", "Confidence", "Lift"]
    col_widths_pp = [35, 35, 16, 12, 14, 10]
    for ci, (h, w) in enumerate(zip(headers_pp, col_widths_pp), 1):
        c           = ws2.cell(row=1, column=ci, value=h)
        c.font      = Font(bold=True, size=10, name="Arial")
        c.border    = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[1].height = 22

    for ri, row in output_pp.iterrows():
        r = ri + 2
        cells = [
            (1, row['Jika Membeli'],        None,      'left'),
            (2, row['Maka Membeli'],        None,      'left'),
            (3, int(row['Jumlah Invoice']), '0',       'center'),
            (4, float(row['Support']),      'General', 'center'),
            (5, float(row['Confidence']),   'General', 'center'),
            (6, float(row['Lift']),         'General', 'center'),
        ]
        for ci, val, fmt, aln in cells:
            c           = ws2.cell(row=r, column=ci, value=val)
            c.font      = Font(size=10, name="Arial")
            c.border    = thin_border()
            c.alignment = Alignment(horizontal=aln)
            if fmt:
                c.number_format = fmt
        ws2.row_dimensions[r].height = 18

    wb.save('retail_insight.xlsx')

    print("\nSelesai! File disimpan:")
    print(f"  retail_insight.xlsx")
    print(f"    - Sheet 'Rising Star'        : {len(rising_stars)} produk")
    print(f"    - Sheet 'Potential Packaging': {len(output_pp)} rules")
    print(f"  rising_star_index.png")
    print(f"  rising_star_actual.png")


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    main()
